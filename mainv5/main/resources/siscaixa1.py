from resources import funcaolimpar as fl
from resources import tela_login_v3 as tl
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3
import os

conexao = sqlite3.connect('hortifruti.db')

# Exibir produtos disponíveis (agora puxando os produtos do banco de dados)
def exibir_produtos(con):
    print("\n--- Produtos Disponíveis ---")
    cursor = con.cursor()
    cursor.execute("SELECT codigo, nome, valor_por_kg, tipo FROM alimentos")
    produtos = cursor.fetchall()
    for produto in produtos:
        codigo, nome, preco, tipo = produto
        print(f"{codigo} - {nome} | R${preco:.2f} ({tipo})")

# Criar e formatar cabeçalho da planilha de vendas
def criar_cabecalho(sheet):
    headers = ["Data/Hora", "Produto", "Quantidade", "Preço Unitário (R$)", "Preço Total (R$)"]
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = 20

# Registrar venda no sistema de caixa
def registrar_venda(con):
    if not os.path.exists("vendas.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        criar_cabecalho(sheet)
        workbook.save("vendas.xlsx")
    else:
        workbook = openpyxl.load_workbook("vendas.xlsx")
        sheet = workbook.active

    exibir_produtos(con)
    total_venda = 0

    while True:
        try:
            codigo_produto = int(input("Digite o código do produto (ou 0 para finalizar): "))
            if codigo_produto == 0:
                break
            
            cursor = con.cursor()
            cursor.execute("SELECT nome, valor_por_kg, tipo FROM alimentos WHERE codigo = ?", (codigo_produto,))
            produto = cursor.fetchone()

            if produto:
                nome, preco_unitario, tipo_venda = produto
                quantidade = float(input(f"Digite a quantidade para {nome} ({tipo_venda}): "))
                preco_total = quantidade * preco_unitario
                total_venda += preco_total
                data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Adicionando linha na planilha
                sheet.append([data_hora, nome, quantidade, preco_unitario, preco_total])
                print(f"{nome} adicionado: {quantidade} x R${preco_unitario:.2f} = R${preco_total:.2f}")
            else:
                print("Código de produto inválido.")
        except ValueError:
            print("Entrada inválida. Por favor, insira um número.")

    workbook.save("vendas.xlsx")
    print(f"\nTotal da venda: R${total_venda:.2f}")

# Menu do sistema de caixa
def loop_menu_caixa():
    con = conexao()
    while True:
        print("\n--- Sistema de Caixa ---")
        print("1 - Registrar venda")
        print("2 - Exibir produtos disponíveis")
        print("3 - Sair")

        opcao = input("Escolha uma opção: ")
        fl.limpar()

        if opcao == "1":
            registrar_venda(con)
        elif opcao == "2":
            exibir_produtos(con)
        elif opcao == "3":
            conexao.close()
            tl.usuario = None
            return
        else:
            print("Opção inválida. Por favor, insira uma opção válida.")
