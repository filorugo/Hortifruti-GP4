# Importações necessárias
import openpyxl
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rich.console import Console
from rich.table import Table
import os

# Console para exibir a tabela Rich
console = Console()

# Dicionário de produtos: código -> (nome, preço por kg ou unidade, tipo de venda)
produtos = {
    1: ("Banana Prata/Nanica", 6.80, "kg"),
    2: ("Abacate", 8.00, "kg"),
    3: ("Manga Tomy", 7.00, "kg"),
    4: ("Coco Seco", 7.00, "unidade"),
    5: ("Goiaba", 8.00, "kg"),
    6: ("Ameixa", 15.00, "kg"),
    7: ("Kiwi", 12.00, "kg"),
    8: ("Tomate Italiano", 8.00, "kg"),
    9: ("Limão Taiti", 5.50, "kg"),
    10: ("Mamão Formosa/Papaya", 6.50, "kg"),
    11: ("Abacaxi", 5.00, "unidade"),
    12: ("Melancia", 4.00, "kg"),
    13: ("Melão Amarelo", 6.00, "kg"),
    14: ("Pepino Japonês", 5.00, "kg"),
    15: ("Maracujá Azedo", 10.00, "kg"),
    16: ("Nectarina", 18.00, "kg"),
    17: ("Pera Willians", 12.00, "kg"),
    18: ("Maçã Argentina", 10.50, "kg"),
    19: ("Abóbora Japonesa", 5.00, "kg"),
    20: ("Batata Inglesa", 5.90, "kg"),
    21: ("Batata Doce", 6.50, "kg"),
    22: ("Batata Baroa", 10.00, "kg"),
    23: ("Berinjela", 6.00, "kg"),
    24: ("Beterraba", 5.00, "kg"),
    25: ("Cebola Nacional", 5.70, "kg"),
    26: ("Cenoura", 6.50, "kg"),
    27: ("Chuchu", 4.50, "kg"),
    28: ("Pimentão Verde", 9.00, "kg"),
    29: ("Gengibre", 18.00, "kg"),
    30: ("Inhame", 8.00, "kg"),
    31: ("Pepino Preto", 4.50, "kg"),
    32: ("Quiabo", 10.00, "kg"),
    33: ("Vagem Americana", 12.00, "kg"),
    34: ("Couve-flor", 6.50, "unidade"),
    35: ("Repolho Verde/Roxo", 5.50, "kg"),
    36: ("Melão Sapo", 8.00, "kg"),
    37: ("Romã", 15.00, "kg"),
    38: ("Pitaya", 20.00, "kg")
}

# Função para exibir a lista de produtos disponíveis
def exibir_produtos():
    table = Table(title="Produtos Disponíveis")
    table.add_column("Código", style="cyan")
    table.add_column("Produto", style="magenta")
    table.add_column("Preço (R$)", justify="right", style="green")
    table.add_column("Venda", style="yellow")

    for codigo, (nome, preco, tipo_venda) in produtos.items():
        table.add_row(str(codigo), nome, f"{preco:.2f}", tipo_venda)

    console.print(table)

# Função para criar e formatar o cabeçalho da planilha
def criar_cabecalho(sheet):
    headers = ["Data/Hora", "Produto", "Quantidade", "Preço Unitário (R$)", "Preço Total (R$)"]
    sheet.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = 20

# Função para estilizar linhas da venda
def estilizar_linhas(sheet, inicio_linha, fim_linha):
    fill_even = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(inicio_linha, fim_linha + 1):
        for col in range(1, 6):  # Colunas A até E
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = fill_even if row % 2 == 0 else fill_odd
            if col in [4, 5]:  # Colunas de preço em formato moeda
                cell.number_format = "R$ #,##0.00"

# Função para calcular o valor do item com base no código e na quantidade (kg ou unidade)
def calcular_valor_produto(codigo, quantidade):
    if codigo in produtos:
        nome_produto, preco, tipo_venda = produtos[codigo]
        valor_item = preco * quantidade
        return nome_produto, preco, quantidade, valor_item
    else:
        return None, 0, 0, 0

# Função para exibir o resumo da venda com Rich
def exibir_resumo_venda(carrinho):
    valor_total = sum(item[3] for item in carrinho)
    
    # Criando a tabela Rich para exibir o resumo
    table = Table(title="Resumo da Venda")
    table.add_column("Produto", style="magenta")
    table.add_column("Quantidade", justify="center")
    table.add_column("Preço Unitário (R$)", justify="right", style="green")
    table.add_column("Preço Total (R$)", justify="right", style="green")
    
    for item in carrinho:
        produto, preco, quantidade, valor_item = item
        table.add_row(produto, f"{quantidade:.2f}", f"{preco:.2f}", f"{valor_item:.2f}")
    
    table.add_row("Total", "", "", f"{valor_total:.2f}", style="bold green")
    
    console.print(table)

# Função para finalizar a venda e salvar na planilha Excel
def finalizar_venda(carrinho):
    valor_total = sum(item[3] for item in carrinho)
    
    # Exibe o resumo da venda com Rich
    exibir_resumo_venda(carrinho)
    
    # Caminho do arquivo Excel
    nome_arquivo = "registro_vendas.xlsx"
    
    # Carregar ou criar a planilha
    if os.path.exists(nome_arquivo):
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active
        inicio_linha = sheet.max_row + 1
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Histórico de Vendas"
        criar_cabecalho(sheet)
        inicio_linha = 2

    # Data e hora da venda
    data_hora_venda = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Adicionar itens da venda
    for item in carrinho:
        produto, preco, quantidade, valor_item = item
        sheet.append([data_hora_venda, produto, quantidade, preco, valor_item])

    # Linha de subtotal
    sheet.append(["", "", "", "Subtotal:", valor_total])
    ultima_linha = sheet.max_row
    for col in range(4, 6):
        cell = sheet.cell(row=ultima_linha, column=col)
        cell.font = Font(bold=True, color="4F81BD")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="double"),
            bottom=Side(style="double"),
        )

    # Estilizar linhas de venda
    estilizar_linhas(sheet, inicio_linha, ultima_linha)

    # Salvar o arquivo
    workbook.save(nome_arquivo)
    print(f"Resumo da venda registrado em {nome_arquivo}")

# Exibir produtos disponíveis
exibir_produtos()

# Carrinho de compras
carrinho = []
total = 0

# Loop para adicionar produtos ao carrinho
while True:
    # Entrada do código do produto com validação
    try:
        codigo = int(input("Digite o código do produto (0 para finalizar): "))
    except ValueError:
        print("Código inválido! Insira um número.")
        continue

    if codigo == 0:
        break
    elif codigo not in produtos:
        print("Código de produto inválido!")
        continue

    # Verifica se o produto é vendido por kg ou unidade
    nome_produto, preco, tipo_venda = produtos[codigo]
    while True:
        try:
            if tipo_venda == "unidade":
                quantidade = float(input(f"Digite a quantidade de {nome_produto}: "))
            else:
                quantidade = float(input(f"Digite o peso em kg de {nome_produto}: "))
            if quantidade <= 0:
                print("A quantidade deve ser maior que zero.")
                continue
            break
        except ValueError:
            print("Quantidade inválida! Insira um número.")

    # Calcula o valor do produto e adiciona ao carrinho
    nome_produto, preco, quantidade, valor_item = calcular_valor_produto(codigo, quantidade)
    print(f"{nome_produto}: R$ {valor_item:.2f}")
    carrinho.append((nome_produto, preco, quantidade, valor_item))
    total += valor_item

# Exibindo o carrinho e total na tabela Rich e registrando na planilha
finalizar_venda(carrinho)
